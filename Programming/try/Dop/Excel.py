import openpyxl

# Создаем новую книгу Excel
wb = openpyxl.Workbook()

# =======================================================================
# Лист "Loan Calculations"
# =======================================================================
ws = wb.active
ws.title = "Loan Calculations"

# Заполняем ячейки
ws["B1"] = 1000000  # Loan amount
ws["B2"] = 120  # Loan term in months (10 years)
ws["B3"] = 0.14  # Interest rate
ws["B4"] = "=PMT(B3/12,B2,-B1)"

ws["E2"] = 180  # Loan term in months (15 years)
ws["E3"] = 0.12  # Interest rate
ws["E4"] = -24500  # Monthly payment
ws["E1"] = "=-PV(E3/12,E2,-E4)"  # Исправленная формула

ws["H1"] = 100000  # Loan amount
ws["H3"] = 0.12  # Interest rate
ws["H4"] = -3000  # Monthly payment
ws["H2"] = "=NPER(H3/12,H4,-H1)"  # Исправленная формула

# =======================================================================
# Лист "Sales Growth"
# =======================================================================

ws2 = wb.create_sheet(title="Sales Growth")

# Исходные данные
ws2["B1"] = 1.2  # Начальный коэффициент роста (можно оставить пустым)
ws2["B2"] = 250000  # Объем продаж за 1-й год

# Годы
for i in range(0, 11):
    ws2["B" + str(6 + i)] = i

# Формулы расчета объема продаж
for i in range(0, 11):
    ws2["C" + str(6 + i)] = "=$B$2*($B$1^B" + str(6 + i) + ")"

# Целевое значение для подбора параметра
target_value = 20000000
ws2["C16"] = target_value  # Объем продаж через 10 лет (целевое значение)


# Подбор параметра для коэффициента роста (Задача 1)
def goal_seek_growth_factor(
    sheet, target_cell, changing_cell, target_value, tolerance=0.001, max_iterations=100
):
    # Итеративный подбор параметра
    for _ in range(max_iterations):
        current_value = sheet[target_cell].value
        if abs(current_value - target_value) <= tolerance:
            break

        # Если текущее значение меньше целевого, увеличиваем коэффициент
        if current_value < target_value:
            sheet[changing_cell].value = sheet[changing_cell].value * 1.01
        else:
            sheet[changing_cell].value = sheet[changing_cell].value * 0.99

    # Явный пересчет формул после подбора параметра
    sheet[target_cell].value = sheet[target_cell].value


goal_seek_growth_factor(ws2, "C16", "B1", target_value)

# Подбор параметра для начального объема продаж (Задача 2)
ws2["B1"] = 1.2  # сбрасываем коэф. роста на начальное значение


def goal_seek_initial_sales(
    sheet, target_cell, changing_cell, target_value, tolerance=0.001, max_iterations=100
):
    for _ in range(max_iterations):
        current_value = sheet[target_cell].value

        if abs(current_value - target_value) <= tolerance:
            break

        # Изменяем начальный объем продаж пропорционально разнице между целевым и текущим значениями
        ratio = target_value / current_value
        sheet[changing_cell].value = sheet[changing_cell].value * ratio

    # Явный пересчет формул после подбора параметра
    sheet[target_cell].value = sheet[target_cell].value


goal_seek_initial_sales(ws2, "C16", "B2", target_value)

# =======================================================================
# Сохранение файла
# =======================================================================
file_path = "Loan_and_Sales_Calculations.xlsx"
wb.save(file_path)

print(f"Файл сохранен как: {file_path}")
